from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

# 定义输入和输出文件路径
input_file_path = r'/Users/msy/Desktop/ranks/mmlu_ranks.xlsx'             # 源文件路径
output_file_path = r'/Users/msy/Desktop/ranks/ranks_processed_borda.xlsx'  # 处理后保存的文件路径

# 加载原始工作簿
try:
    wb = load_workbook(input_file_path)
except FileNotFoundError:
    print(f"文件 {input_file_path} 未找到。请检查路径是否正确。")
    exit(1)
except Exception as e:
    print(f"加载工作簿时出错: {e}")
    exit(1)

# 创建一个新的工作簿用于保存处理后的数据
# 这里我们复制原始工作簿，以保留原始格式和样式
from copy import copy

wb_processed = load_workbook(input_file_path)

# 遍历每个工作表
for sheet_name in wb.sheetnames:
    print(f"处理工作表: {sheet_name}")
    ws = wb[sheet_name]
    ws_processed = wb_processed[sheet_name]
    
    # 读取数据：假设第一行是标题（候选人姓名），后续行是评委的排名
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    if len(data) < 2:
        print(f"工作表 '{sheet_name}' 数据不足，跳过。")
        continue
    
    headers = data[0]
    rankings = data[1:]
    
    # 确保每个候选人都有一个名称
    if any(header is None for header in headers):
        print(f"工作表 '{sheet_name}' 存在空的候选人名称，跳过。")
        continue
    
    candidates = headers
    num_candidates = len(candidates)
    
    # 初始化总分字典
    total_scores = {candidate: 0 for candidate in candidates}
    
    # 计算总分
    for judge_index, row in enumerate(rankings, start=2):  # 从第2行开始（第1行为标题）
        for col_index, rank in enumerate(row):
            candidate = candidates[col_index]
            if rank is None:
                continue  # 跳过空值
            try:
                rank = int(rank)
                if 1 <= rank <= num_candidates:
                    total_scores[candidate] += (num_candidates - rank)
                else:
                    print(f"工作表 '{sheet_name}' 的第 {judge_index} 行，候选人 '{candidate}' 的排名 {rank} 超出范围。")
            except ValueError:
                print(f"数据转换错误：在工作表 '{sheet_name}' 的第 {judge_index} 行，第 {col_index + 1} 列无法将 '{rank}' 转换为整数。")
    
    # 根据总分排序候选人
    sorted_candidates = sorted(total_scores.items(), key=lambda x: x[1], reverse=True)
    final_ranking = [candidate for candidate, score in sorted_candidates]
    final_scores = [score for candidate, score in sorted_candidates]
    
    # 写入总分和排名到指定的行
    # 假设第9行用于写入总分，第10行用于写入排名
    scores_row = 9
    ranking_row = 10
    
    # 确保工作表有足够的行
    max_row = max(scores_row, ranking_row)
    if ws_processed.max_row < max_row:
        ws_processed.insert_rows(ws_processed.max_row + 1, amount=(max_row - ws_processed.max_row))
    
    # 写入总分
    for col_num, score in enumerate(final_scores, start=1):
        cell = ws_processed.cell(row=scores_row, column=col_num)
        cell.value = score
    
    # 写入排名
    for col_num, candidate in enumerate(final_ranking, start=1):
        cell = ws_processed.cell(row=ranking_row, column=col_num)
        cell.value = candidate
    
    print(f"工作表 '{sheet_name}' 处理完成。")

# 保存处理后的工作簿
try:
    wb_processed.save(output_file_path)
    print(f"所有工作表已成功处理并保存到 '{output_file_path}'。")
except Exception as e:
    print(f"保存工作簿时出错: {e}")








