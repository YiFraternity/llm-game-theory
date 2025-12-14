"""
计算并导出不同基准（gpqa、gsm8k_cot、mmlu）中模型排名的Borda分数与CHATBOT_ARENA_RANK（预设的模型排名）的皮尔逊相关性。

1. 计算每个模型的Borda分数，并计算其与CHATBOT_ARENA_RANK（预设的模型排名）的皮尔逊相关系数。
2. 将所有基准测试的皮尔逊相关系数整理成一个DataFrame，并导出为Excel文件。导出时：
    - 每个基准测试占一列
    - 每行代表一个问题的皮尔逊相关系数
    - 最后一行是每个基准测试所有问题的平均皮尔逊相关系数
    - 平均值与问题结果之间用5行空值隔开，便于区分
3. 最终输出的Excel文件（average_pearson_correlations.xlsx）可以直观地比较不同基准测试中模型排名的相关性分布。

"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 定义输入和输出文件路径
input_file_path = r'lm-evaluation-harness/output/gpqa_ranks.xlsx'
output_file_path = r'lm-evaluation-harness/output/gpqa_ranks_borda.xlsx'

def borda(file_path, output_path):
    try:
        wb = load_workbook(file_path)
        wb_processed = load_workbook(file_path)  # 保留原始格式的副本
    except Exception as e:
        print(f"文件加载失败: {e}")
        exit(1)

    for sheet_name in wb.sheetnames:
        print(f"\n处理工作表: {sheet_name}")

        # 读取原始数据
        ws = wb[sheet_name]
        data = list(ws.values)

        if len(data) < 2:
            print("数据不足，跳过")
            continue

        headers = data[0]
        rankings = data[1:]

        # 生成候选人编号映射 (列位置1-based)
        candidate_ids = [i+1 for i in range(len(headers))]


        # 按分数排序（分数相同按原始列顺序）
        sorted_ranking = sorted(borda_score(rankings).items(),
                            key=lambda x: (-x[1], x[0]))  # 先按分数降序，再按列号升序

        # 提取排名的列号序列
        final_order = [item[0] for item in sorted_ranking]

        # 写入处理后的工作表
        ws_proc = wb_processed[sheet_name]

        # 在第9行写入分数（保持原始代码逻辑）
        for col_num, (cid, score) in enumerate(sorted_ranking, start=1):
            ws_proc.cell(row=9, column=col_num, value=score)

        # 在第10行写入列号排名（关键修改部分）
        for col_num, cid in enumerate(final_order, start=1):
            ws_proc.cell(row=10, column=col_num, value=cid)

        print(f"处理完成，排名顺序: {final_order}")

    try:
        wb_processed.save(output_file_path)
        print(f"\n文件已保存至: {output_file_path}")
    except Exception as e:
        print(f"保存失败: {e}")

borda(input_file_path, output_file_path)
